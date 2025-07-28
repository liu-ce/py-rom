# accounts_reader.py
from openpyxl import load_workbook
import os

PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))



def load_accounts(excel_path):
    full_path = os.path.join(PROJECT_ROOT, excel_path)
    wb = load_workbook(full_path, read_only=True, data_only=True)
    ws = wb.active  # 默认第一张表
    accounts = []
    row_index = 2  # 从第2行开始（跳过表头）
    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        email, password, recovery = row[:3]  # 只取前3列
        accounts.append({
            "email": str(email),
            "password": str(password),
            "recovery": str(recovery),
            "row_index": row_index,  # 记录行号，用于后续更新
        })
        row_index += 1
    wb.close()
    return accounts


def update_account_status(excel_path, row_index, status="完成"):
    """
    更新账号状态到Excel第四列
    :param excel_path: Excel文件路径
    :param row_index: 行号（从2开始）
    :param status: 状态文字，默认"完成"
    """
    full_path = os.path.join(PROJECT_ROOT, excel_path)
    wb = load_workbook(full_path)
    ws = wb.active
    
    # 更新第四列（D列）
    ws.cell(row=row_index, column=4, value=status)
    
    wb.save(full_path)
    wb.close()
    print(f"✅ 已更新Excel行 {row_index}，状态: {status}")
